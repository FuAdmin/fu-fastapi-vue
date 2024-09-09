import os
from datetime import datetime
import openpyxl
from config import STATIC_URL


def excel_to_dict(file_path, label_field_dict):
    # 加载工作簿
    workbook = openpyxl.load_workbook(filename=file_path)
    sheet = workbook.active  # 获取活动表单

    # 获取第一行作为键
    keys = [label_field_dict.get(cell.value) for cell in sheet[1]]

    # 遍历其他行作为值
    data_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data_list.append(dict(zip(keys, row)))

    return data_list


def dict_to_excel(list_data: list[dict]):
    # 创建并初始化Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    # 向Excel中写入数据
    for index, data in enumerate(list_data):
        if index == 0:
            # 写入表头
            ws.append(list(data.keys()))
        row = [
            ",".join(value) if type(value) is list else value
            for value in list(data.values())
        ]
        ws.append(row)

    # 生成唯一的文件名并设置文件路径
    file_name = datetime.now().strftime("%Y%m%d%H%M%S%f") + ".xlsx"
    current_ymd = datetime.now().strftime("%Y%m%d")
    file_path = os.path.join(STATIC_URL, current_ymd)
    file_url = os.path.join(file_path, file_name)
    # 确保导出文件的目录存在
    if not os.path.exists(file_path):
        os.makedirs(file_path)

    # 保存Excel文件到指定路径
    wb.save(file_url)
    wb.close()
    return file_url
